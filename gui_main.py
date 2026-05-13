"""
气体传感器实验数据分析 - 图形界面版本。
使用 PySide6 构建 GUI，集成 Matplotlib 绘图。
"""

import sys
import io
from pathlib import Path
from typing import Optional, Tuple, List

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QPushButton, QFileDialog,
    QDialog, QFormLayout, QSpinBox, QDoubleSpinBox, QCheckBox,
    QDialogButtonBox, QMessageBox, QHeaderView, QLabel, QGroupBox,
    QSplitter, QSizePolicy, QFrame, QComboBox, QColorDialog,
    QLineEdit, QProgressDialog
)
from PySide6.QtCore import Qt, QMimeData
from PySide6.QtGui import QAction, QIcon, QClipboard, QFont, QColor

import matplotlib
matplotlib.use('QtAgg')
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.lines import Line2D

from settings_manager import SettingsManager, AppSettings, AnalysisSettings
from analyze_core import run_analysis, AnalysisResult
from language_manager import LanguageManager


# Windows console UTF-8 output
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


# 蓝白配色扁平化样式表
FLAT_STYLE = """
QMainWindow, QDialog {
    background-color: #f5f7fa;
}
QWidget {
    font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif;
    font-size: 11px;
}
QGroupBox {
    background-color: white;
    border: 1px solid #d0d7de;
    border-radius: 6px;
    margin-top: 10px;
    padding-top: 15px;
    font-weight: bold;
    color: #24292f;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 10px;
    padding: 0 5px;
}
QPushButton {
    background-color: #0969da;
    color: white;
    border: none;
    border-radius: 6px;
    padding: 8px 16px;
    font-weight: bold;
    min-width: 80px;
}
QPushButton:hover {
    background-color: #0860ca;
}
QPushButton:pressed {
    background-color: #0757b5;
}
QPushButton:disabled {
    background-color: #8c959f;
}
QTableWidget {
    background-color: white;
    border: 1px solid #d0d7de;
    border-radius: 6px;
    gridline-color: #e1e4e8;
    selection-background-color: #ddf4ff;
    selection-color: #24292f;
}
QTableWidget::item {
    padding: 5px;
}
QHeaderView::section {
    background-color: #f6f8fa;
    border: none;
    border-bottom: 2px solid #0969da;
    padding: 8px;
    font-weight: bold;
    color: #24292f;
}
QCheckBox {
    spacing: 8px;
    color: #24292f;
}
QCheckBox::indicator {
    width: 18px;
    height: 18px;
    border: 2px solid #d0d7de;
    border-radius: 4px;
    background-color: white;
}
QCheckBox::indicator:checked {
    background-color: #0969da;
    border-color: #0969da;
}
QCheckBox::indicator:disabled {
    background-color: #f6f8fa;
    border-color: #e1e4e8;
}
QRadioButton {
    spacing: 8px;
    color: #24292f;
}
QRadioButton::indicator {
    width: 18px;
    height: 18px;
    border: 2px solid #d0d7de;
    border-radius: 9px;
    background-color: white;
}
QRadioButton::indicator:checked {
    background-color: #0969da;
    border-color: #0969da;
}
QComboBox {
    border: 1px solid #d0d7de;
    border-radius: 6px;
    padding: 5px 10px;
    background-color: white;
    color: #24292f;
}
QComboBox:hover {
    border-color: #0969da;
}
QComboBox::drop-down {
    border: none;
}
QLabel {
    color: #24292f;
}
QLineEdit {
    border: 1px solid #d0d7de;
    border-radius: 4px;
    padding: 5px 8px;
    background-color: white;
}
QLineEdit:hover {
    border-color: #0969da;
}
QMenuBar {
    background-color: white;
    border-bottom: 1px solid #d0d7de;
}
QMenuBar::item:selected {
    background-color: #ddf4ff;
}
QMenu {
    background-color: white;
    border: 1px solid #d0d7de;
}
QMenu::item:selected {
    background-color: #ddf4ff;
}
QStatusBar {
    background-color: #f6f8fa;
    border-top: 1px solid #d0d7de;
}
QSplitter::handle {
    background-color: #d0d7de;
}
QSpinBox, QDoubleSpinBox {
    border: 1px solid #d0d7de;
    border-radius: 4px;
    padding: 4px;
    background-color: white;
}
QSpinBox:hover, QDoubleSpinBox:hover {
    border-color: #0969da;
}
QSpinBox::up-button, QDoubleSpinBox::up-button {
    subcontrol-origin: border;
    subcontrol-position: top right;
    border-left: 1px solid #d0d7de;
    border-bottom: 1px solid #d0d7de;
    border-top-right-radius: 4px;
    background-color: #f6f8fa;
    width: 20px;
}
QSpinBox::up-button:hover, QDoubleSpinBox::up-button:hover {
    background-color: #ddf4ff;
}
QSpinBox::down-button, QDoubleSpinBox::down-button {
    subcontrol-origin: border;
    subcontrol-position: bottom right;
    border-left: 1px solid #d0d7de;
    border-top: 1px solid #d0d7de;
    border-bottom-right-radius: 4px;
    background-color: #f6f8fa;
    width: 20px;
}
QSpinBox::down-button:hover, QDoubleSpinBox::down-button:hover {
    background-color: #ddf4ff;
}
QSpinBox::up-arrow, QDoubleSpinBox::up-arrow {
    width: 0;
    height: 0;
    border-left: 5px solid transparent;
    border-right: 5px solid transparent;
    border-bottom: 7px solid #24292f;
}
QSpinBox::down-arrow, QDoubleSpinBox::down-arrow {
    width: 0;
    height: 0;
    border-left: 5px solid transparent;
    border-right: 5px solid transparent;
    border-top: 7px solid #24292f;
}
QProgressDialog {
    background-color: white;
}
"""


def get_current_unit_and_scale(max_val: float) -> Tuple[str, float]:
    """
    根据电流最大值自动选择合适的单位和缩放因子。
    返回 (单位字符串, 缩放因子)。
    规则：当最大值 > 5000X 时，升级到下一个单位。
    """
    abs_max = abs(max_val)
    if abs_max == 0:
        return 'A', 1.0

    if abs_max < 1e-9:
        return 'pA', 1e12
    elif abs_max < 1e-6:
        return 'nA', 1e9
    elif abs_max < 1e-3:
        return 'μA', 1e6
    elif abs_max < 1.0:
        return 'mA', 1e3
    else:
        return 'A', 1.0


def get_voltage_unit_and_scale(max_val: float) -> Tuple[str, float]:
    """
    根据电压最大值自动选择合适的单位和缩放因子。
    返回 (单位字符串, 缩放因子)。
    """
    abs_max = abs(max_val)
    if abs_max == 0:
        return 'V', 1.0

    if abs_max < 1e-3:
        return 'mV', 1e3
    elif abs_max < 1.0:
        return 'V', 1.0
    elif abs_max < 1e3:
        return 'kV', 1e-3
    else:
        return 'MV', 1e-6


def get_resistance_unit_and_scale(max_val: float) -> Tuple[str, float]:
    """
    根据电阻最大值自动选择合适的单位和缩放因子。
    返回 (单位字符串, 缩放因子)。
    """
    abs_max = abs(max_val)
    if abs_max == 0:
        return 'Ω', 1.0

    if abs_max < 1e3:
        return 'Ω', 1.0
    elif abs_max < 1e6:
        return 'kΩ', 1e-3
    elif abs_max < 1e9:
        return 'MΩ', 1e-6
    else:
        return 'GΩ', 1e-9


def _column_unit_scale(values: List[float], unit_func) -> Tuple[str, float]:
    """根据一列中的有效值确定该列的单位和缩放因子。"""
    valid = [v for v in values if v is not None and np.isfinite(v) and v > 0]
    if not valid:
        return unit_func(0)[0], 1.0
    return unit_func(max(valid))


def _make_form_group(title: str, rows: list, extra_widget=None):
    """创建表单分组。rows >= 2 时使用双列布局。"""
    group = QGroupBox(title)
    if len(rows) < 2:
        form = QFormLayout()
        for label_text, widget in rows:
            form.addRow(label_text, widget)
        if extra_widget:
            form.addRow("", extra_widget)
        group.setLayout(form)
    else:
        outer = QHBoxLayout()
        mid = (len(rows) + 1) // 2
        for col_idx, col_start in enumerate((0, mid)):
            form = QFormLayout()
            for i in range(col_start, min(col_start + mid, len(rows))):
                label_text, widget = rows[i]
                form.addRow(label_text, widget)
            if extra_widget and col_idx == 1:
                form.addRow("", extra_widget)
            outer.addLayout(form)
        group.setLayout(outer)
    return group


class ColorButton(QPushButton):
    """点击弹出颜色选择对话框的小按钮"""

    def __init__(self, hex_color: str, parent=None):
        super().__init__(parent)
        self._color = hex_color
        self.setFixedSize(36, 22)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.clicked.connect(self._pick)
        self._update_style()

    def _update_style(self):
        self.setStyleSheet(
            f"QPushButton{{background-color:{self._color}; border:1px solid #b0b7c0; "
            f"border-radius:3px; min-width:36px; min-height:22px;}}"
            f"QPushButton:hover{{border-color:#0969da;}}"
        )

    def _pick(self):
        color = QColorDialog.getColor(QColor(self._color), self)
        if color.isValid():
            self._color = color.name()
            self._update_style()

    def color(self) -> str:
        return self._color


class SystemSettingsDialog(QDialog):
    """系统首选项对话框：分析参数、峰值检测、界面缩放、语言"""

    def __init__(self, settings: AppSettings, lang: LanguageManager, parent=None):
        super().__init__(parent)
        self.settings = settings
        self.lang = lang
        self.setWindowTitle(self.lang.get("dialog_system_settings"))
        self.setMinimumWidth(400)
        self._setup_ui()

    def _tr(self, key: str, **kwargs) -> str:
        return self.lang.get(key, **kwargs)

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        self.rise_lower = QDoubleSpinBox()
        self.rise_lower.setRange(0.1, 99.9)
        self.rise_lower.setValue(self.settings.analysis.rise_lower_percent)
        self.rise_lower.setSuffix(" %")

        self.rise_upper = QDoubleSpinBox()
        self.rise_upper.setRange(0.1, 99.9)
        self.rise_upper.setValue(self.settings.analysis.rise_upper_percent)
        self.rise_upper.setSuffix(" %")

        self.fall_upper = QDoubleSpinBox()
        self.fall_upper.setRange(0.1, 99.9)
        self.fall_upper.setValue(self.settings.analysis.fall_upper_percent)
        self.fall_upper.setSuffix(" %")

        self.fall_lower = QDoubleSpinBox()
        self.fall_lower.setRange(0.1, 99.9)
        self.fall_lower.setValue(self.settings.analysis.fall_lower_percent)
        self.fall_lower.setSuffix(" %")

        analysis_rows = [
            (self._tr("label_rise_lower"), self.rise_lower),
            (self._tr("label_rise_upper"), self.rise_upper),
            (self._tr("label_fall_upper"), self.fall_upper),
            (self._tr("label_fall_lower"), self.fall_lower),
        ]
        layout.addWidget(_make_form_group(self._tr("group_analysis"), analysis_rows))

        self.prominence = QDoubleSpinBox()
        self.prominence.setRange(0.01, 5.0)
        self.prominence.setValue(self.settings.analysis.prominence)
        self.prominence.setSingleStep(0.05)
        self.prominence.setDecimals(2)

        self.distance = QSpinBox()
        self.distance.setRange(10, 1000)
        self.distance.setValue(self.settings.analysis.distance)
        self.distance.setSuffix(" pts")

        self.baseline_fraction = QDoubleSpinBox()
        self.baseline_fraction.setRange(0.001, 0.5)
        self.baseline_fraction.setValue(self.settings.analysis.baseline_fraction)
        self.baseline_fraction.setSingleStep(0.005)
        self.baseline_fraction.setDecimals(3)

        self.peak_threshold = QDoubleSpinBox()
        self.peak_threshold.setRange(1e-12, 1e-3)
        self.peak_threshold.setValue(self.settings.analysis.peak_threshold)
        self.peak_threshold.setSingleStep(1e-8)
        self.peak_threshold.setDecimals(10)
        self.peak_threshold.setSuffix(" A")

        reset_btn = QPushButton(self._tr("btn_reset"))
        reset_btn.setStyleSheet("""
            QPushButton {
                background-color: #6e7781;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #57606a;
            }
        """)
        reset_btn.clicked.connect(self._reset_detection)

        detection_rows = [
            (self._tr("label_prominence"), self.prominence),
            (self._tr("label_distance"), self.distance),
            (self._tr("label_baseline"), self.baseline_fraction),
            (self._tr("label_peak_thresh"), self.peak_threshold),
        ]
        layout.addWidget(_make_form_group(self._tr("group_detection"), detection_rows, reset_btn))

        self.gui_scale = QDoubleSpinBox()
        self.gui_scale.setRange(0.5, 3.0)
        self.gui_scale.setValue(self.settings.gui_scale)
        self.gui_scale.setSingleStep(0.1)
        self.gui_scale.setSuffix(" x")

        scale_rows = [(self._tr("label_gui_scale"), self.gui_scale)]
        layout.addWidget(_make_form_group(self._tr("group_gui_scale"), scale_rows))

        self.combo_language = QComboBox()
        for code in self.lang.get_available_languages():
            self.combo_language.addItem(self.lang.get_language_name(code), code)
        current_idx = self.lang.get_available_languages().index(self.lang.language)
        self.combo_language.setCurrentIndex(current_idx)

        lang_rows = [(self._tr("label_language"), self.combo_language)]
        layout.addWidget(_make_form_group(self._tr("group_language"), lang_rows))

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def _reset_detection(self):
        """重置峰值检测参数为默认值"""
        defaults = AnalysisSettings()
        self.prominence.setValue(defaults.prominence)
        self.distance.setValue(defaults.distance)
        self.baseline_fraction.setValue(defaults.baseline_fraction)
        self.peak_threshold.setValue(defaults.peak_threshold)

    def get_settings(self) -> dict:
        return {
            'analysis': {
                'rise_lower_percent': self.rise_lower.value(),
                'rise_upper_percent': self.rise_upper.value(),
                'fall_upper_percent': self.fall_upper.value(),
                'fall_lower_percent': self.fall_lower.value(),
                'prominence': self.prominence.value(),
                'distance': self.distance.value(),
                'baseline_fraction': self.baseline_fraction.value(),
                'peak_threshold': self.peak_threshold.value(),
            },
            'gui_scale': self.gui_scale.value(),
            'language': self.combo_language.currentData(),
        }

    def get_analysis_changed(self) -> bool:
        return (
            self.rise_lower.value() != self.settings.analysis.rise_lower_percent or
            self.rise_upper.value() != self.settings.analysis.rise_upper_percent or
            self.fall_upper.value() != self.settings.analysis.fall_upper_percent or
            self.fall_lower.value() != self.settings.analysis.fall_lower_percent or
            self.prominence.value() != self.settings.analysis.prominence or
            self.distance.value() != self.settings.analysis.distance or
            self.baseline_fraction.value() != self.settings.analysis.baseline_fraction or
            self.peak_threshold.value() != self.settings.analysis.peak_threshold
        )


class PlotSettingsDialog(QDialog):
    """绘图首选项对话框：字体、线宽、颜色"""

    def __init__(self, settings: AppSettings, lang: LanguageManager, parent=None):
        super().__init__(parent)
        self.settings = settings
        self.lang = lang
        self.setWindowTitle(self.lang.get("dialog_plot_settings"))
        self.setMinimumWidth(400)
        self._setup_ui()

    def _tr(self, key: str, **kwargs) -> str:
        return self.lang.get(key, **kwargs)

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        plot_settings = self.settings.plot

        self.title_size = QSpinBox()
        self.title_size.setRange(8, 36)
        self.title_size.setValue(plot_settings.font_size_title)

        self.title_bold = QCheckBox()
        self.title_bold.setChecked(plot_settings.title_bold)

        self.axis_label_size = QSpinBox()
        self.axis_label_size.setRange(8, 24)
        self.axis_label_size.setValue(plot_settings.font_size_axis_label)

        self.axis_label_bold = QCheckBox()
        self.axis_label_bold.setChecked(plot_settings.axis_label_bold)

        self.tick_size = QSpinBox()
        self.tick_size.setRange(6, 18)
        self.tick_size.setValue(plot_settings.font_size_tick)

        self.legend_size = QSpinBox()
        self.legend_size.setRange(6, 18)
        self.legend_size.setValue(plot_settings.font_size_legend)

        self.annotation_size = QSpinBox()
        self.annotation_size.setRange(6, 16)
        self.annotation_size.setValue(plot_settings.font_size_annotation)

        self.line_width_raw = QDoubleSpinBox()
        self.line_width_raw.setRange(0.1, 5.0)
        self.line_width_raw.setValue(plot_settings.line_width_raw)
        self.line_width_raw.setSingleStep(0.1)

        self.line_width_smooth = QDoubleSpinBox()
        self.line_width_smooth.setRange(0.1, 5.0)
        self.line_width_smooth.setValue(plot_settings.line_width_smooth)
        self.line_width_smooth.setSingleStep(0.1)

        self.line_width_edge = QDoubleSpinBox()
        self.line_width_edge.setRange(0.5, 8.0)
        self.line_width_edge.setValue(plot_settings.line_width_edge)
        self.line_width_edge.setSingleStep(0.5)

        self.btn_color_raw = ColorButton(plot_settings.color_raw)
        self.btn_color_smooth = ColorButton(plot_settings.color_smooth)
        self.btn_color_rise = ColorButton(plot_settings.color_rise)
        self.btn_color_fall = ColorButton(plot_settings.color_fall)
        self.btn_color_peak = ColorButton(plot_settings.color_peak)
        self.btn_color_rise_marker = ColorButton(plot_settings.color_rise_marker)
        self.btn_color_fall_marker = ColorButton(plot_settings.color_fall_marker)
        self.btn_color_voltage = ColorButton(plot_settings.color_voltage)
        self.btn_color_resistance = ColorButton(plot_settings.color_resistance)

        plot_rows = [
            (self._tr("label_title_size"), self.title_size),
            (self._tr("label_title_bold"), self.title_bold),
            (self._tr("label_axis_size"), self.axis_label_size),
            (self._tr("label_axis_bold"), self.axis_label_bold),
            (self._tr("label_tick_size"), self.tick_size),
            (self._tr("label_legend_size"), self.legend_size),
            (self._tr("label_annot_size"), self.annotation_size),
            (self._tr("label_line_width_raw"), self.line_width_raw),
            (self._tr("label_line_width_smooth"), self.line_width_smooth),
            (self._tr("label_line_width_edge"), self.line_width_edge),
            (self._tr("label_color_raw"), self.btn_color_raw),
            (self._tr("label_color_smooth"), self.btn_color_smooth),
            (self._tr("label_color_rise"), self.btn_color_rise),
            (self._tr("label_color_fall"), self.btn_color_fall),
            (self._tr("label_color_peak"), self.btn_color_peak),
            (self._tr("label_color_rise_marker"), self.btn_color_rise_marker),
            (self._tr("label_color_fall_marker"), self.btn_color_fall_marker),
            (self._tr("label_color_voltage"), self.btn_color_voltage),
            (self._tr("label_color_resistance"), self.btn_color_resistance),
        ]
        layout.addWidget(_make_form_group(self._tr("group_plot"), plot_rows))

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def get_settings(self) -> dict:
        return {
            'plot': {
                'font_size_title': self.title_size.value(),
                'title_bold': self.title_bold.isChecked(),
                'font_size_axis_label': self.axis_label_size.value(),
                'axis_label_bold': self.axis_label_bold.isChecked(),
                'font_size_tick': self.tick_size.value(),
                'font_size_legend': self.legend_size.value(),
                'font_size_annotation': self.annotation_size.value(),
                'line_width_raw': self.line_width_raw.value(),
                'line_width_smooth': self.line_width_smooth.value(),
                'line_width_edge': self.line_width_edge.value(),
                'color_raw': self.btn_color_raw.color(),
                'color_smooth': self.btn_color_smooth.color(),
                'color_rise': self.btn_color_rise.color(),
                'color_fall': self.btn_color_fall.color(),
                'color_peak': self.btn_color_peak.color(),
                'color_rise_marker': self.btn_color_rise_marker.color(),
                'color_fall_marker': self.btn_color_fall_marker.color(),
                'color_voltage': self.btn_color_voltage.color(),
                'color_resistance': self.btn_color_resistance.color(),
            },
        }


class BatchConfirmDialog(QDialog):
    """批量处理确认对话框：显示当前分析设置摘要"""

    def __init__(self, settings: AppSettings, file_count: int,
                 lang: LanguageManager, parent=None):
        super().__init__(parent)
        self.settings = settings
        self.file_count = file_count
        self.lang = lang
        self.setWindowTitle(self._tr("dlg_batch_confirm_title"))
        self.setMinimumWidth(400)
        self._setup_ui()

    def _tr(self, key: str, **kwargs) -> str:
        return self.lang.get(key, **kwargs)

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        # 摘要文本
        summary = QLabel(
            self._tr("msg_batch_summary", n=self.file_count)
        )
        summary.setWordWrap(True)
        summary.setStyleSheet("font-size: 12px; font-weight: bold; margin-bottom: 10px;")
        layout.addWidget(summary)

        a = self.settings.analysis

        # 通道A 设置摘要
        ch_a_rows = [
            (self._tr("label_rise_lower"), QLabel(f"{a.rise_lower_percent:.1f} %")),
            (self._tr("label_rise_upper"), QLabel(f"{a.rise_upper_percent:.1f} %")),
            (self._tr("label_fall_upper"), QLabel(f"{a.fall_upper_percent:.1f} %")),
            (self._tr("label_fall_lower"), QLabel(f"{a.fall_lower_percent:.1f} %")),
            (self._tr("label_prominence"), QLabel(f"{a.prominence:.2f}")),
            (self._tr("label_distance"), QLabel(f"{a.distance} pts")),
            (self._tr("label_baseline"), QLabel(f"{a.baseline_fraction:.3f}")),
            (self._tr("label_peak_thresh"), QLabel(f"{a.peak_threshold:.2e} A")),
        ]
        layout.addWidget(_make_form_group(
            self._tr("group_ch_a_summary"), ch_a_rows
        ))

        # 通道B 设置摘要
        b_name = self.settings.b_channel
        if b_name == "resistance":
            b_display = self._tr("opt_b_resistance")
        elif b_name == "voltage":
            b_display = self._tr("opt_b_voltage")
        else:
            b_display = self._tr("opt_b_none")
        ch_b_rows = [(self._tr("label_channel_b"), QLabel(b_display))]
        layout.addWidget(_make_form_group(
            self._tr("group_ch_b_summary"), ch_b_rows
        ))

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)


class BatchOutputDialog(QDialog):
    """批量输出设置对话框：Excel名称、图像类型、输出目录"""

    def __init__(self, lang: LanguageManager, parent=None):
        super().__init__(parent)
        self.lang = lang
        self.setWindowTitle(self._tr("dlg_batch_output_title"))
        self.setMinimumWidth(450)
        self._setup_ui()

    def _tr(self, key: str, **kwargs) -> str:
        return self.lang.get(key, **kwargs)

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        # 汇总表名称
        excel_group = QGroupBox(self._tr("group_excel_name"))
        excel_layout = QVBoxLayout(excel_group)
        self.edit_excel_name = QLineEdit("batch_results.xlsx")
        excel_layout.addWidget(self.edit_excel_name)
        layout.addWidget(excel_group)

        # 图像类型
        img_group = QGroupBox(self._tr("group_image_types"))
        img_layout = QVBoxLayout(img_group)
        self.chk_svg = QCheckBox(self._tr("chk_svg"))
        self.chk_svg.setChecked(True)
        img_layout.addWidget(self.chk_svg)
        self.chk_png = QCheckBox(self._tr("chk_png"))
        self.chk_png.setChecked(True)
        img_layout.addWidget(self.chk_png)
        layout.addWidget(img_group)

        # 输出目录
        dir_group = QGroupBox(self._tr("group_output_dir"))
        dir_layout = QHBoxLayout(dir_group)
        self.edit_output_dir = QLineEdit(str(Path.home() / "Desktop"))
        dir_layout.addWidget(self.edit_output_dir)
        btn_select = QPushButton(self._tr("btn_select_dir"))
        btn_select.clicked.connect(self._select_dir)
        dir_layout.addWidget(btn_select)
        layout.addWidget(dir_group)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def _select_dir(self):
        d = QFileDialog.getExistingDirectory(
            self, self._tr("dlg_select_dir_title"), self.edit_output_dir.text()
        )
        if d:
            self.edit_output_dir.setText(d)

    def get_excel_name(self) -> str:
        return self.edit_output_dir.text().strip()

    def get_output_dir(self) -> str:
        return self.edit_output_dir.text().strip()

    def get_svg_enabled(self) -> bool:
        return self.chk_svg.isChecked()

    def get_png_enabled(self) -> bool:
        return self.chk_png.isChecked()


class PlotCanvas(FigureCanvas):
    """Matplotlib 绘图画布"""

    def __init__(self, parent=None):
        self.fig = Figure(figsize=(12, 6), dpi=100)
        super().__init__(self.fig)
        self.setParent(parent)
        self.ax = self.fig.add_subplot(111)
        self.ax_right = None
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def plot_data(self, result: AnalysisResult, settings: AppSettings,
                  b_channel: str = "",
                  left_scale: str = 'linear', right_scale: str = 'linear'):
        """绘制分析结果。"""
        self.fig.clear()

        times = result.times
        currents = result.currents
        voltages = result.voltages
        resistances = result.resistances
        currents_smooth = result.currents_smooth
        rounds = result.rounds
        plot_settings = settings.plot
        analysis_settings = settings.analysis

        current_unit, current_scale = get_current_unit_and_scale(np.max(currents))
        font_family = 'Times New Roman'

        currents_scaled = currents * current_scale
        currents_smooth_scaled = currents_smooth * current_scale

        self.ax = self.fig.add_subplot(111)

        self.ax.plot(
            times, currents_scaled,
            color=plot_settings.color_raw, linewidth=plot_settings.line_width_raw, alpha=0.4,
            label='Raw data'
        )
        self.ax.plot(
            times, currents_smooth_scaled,
            color=plot_settings.color_smooth, linewidth=plot_settings.line_width_smooth, alpha=0.7,
            label='Smoothed'
        )

        for r in rounds:
            start, peak, end = r.start_idx, r.peak_idx, r.end_idx

            self.ax.plot(
                times[start:peak + 1], currents_smooth_scaled[start:peak + 1],
                color=plot_settings.color_rise, linewidth=plot_settings.line_width_edge,
            )
            self.ax.plot(
                times[peak:end + 1], currents_smooth_scaled[peak:end + 1],
                color=plot_settings.color_fall, linewidth=plot_settings.line_width_edge,
            )

            self.ax.plot(times[peak], currents_smooth_scaled[peak], '*',
                         color=plot_settings.color_peak, markersize=10)

            self.ax.annotate(
                f'#{r.round_num}',
                xy=(times[peak], currents_smooth_scaled[peak]),
                xytext=(5, 10), textcoords='offset points',
                fontsize=plot_settings.font_size_annotation,
                color='darkred', fontfamily=font_family,
            )

            if r.rise:
                for key, marker in [('t_lower', 'o'), ('t_upper', 's')]:
                    t_val = getattr(r.rise, key)
                    idx = np.argmin(np.abs(times - t_val))
                    self.ax.plot(t_val, currents_smooth_scaled[idx], marker,
                                 color=plot_settings.color_rise_marker, markersize=5)
            if r.fall:
                for key, marker in [('t_upper', 'o'), ('t_lower', 's')]:
                    t_val = getattr(r.fall, key)
                    idx = np.argmin(np.abs(times - t_val))
                    self.ax.plot(t_val, currents_smooth_scaled[idx], marker,
                                 color=plot_settings.color_fall_marker, markersize=5)

        self.ax.set_ylabel(
            f'Current ({current_unit})',
            fontsize=plot_settings.font_size_axis_label,
            fontweight='bold' if plot_settings.axis_label_bold else 'normal',
            fontfamily=font_family,
        )
        self.ax.set_yscale(left_scale)

        self.ax_right = None
        if b_channel == "voltage":
            self.ax_right = self.ax.twinx()
            self.ax_right.set_zorder(self.ax.get_zorder() - 1)
            self.ax.patch.set_visible(False)

            voltage_unit, voltage_scale = get_voltage_unit_and_scale(np.max(voltages))
            voltages_scaled = voltages * voltage_scale

            voltage_rms = np.sqrt(np.mean(voltages**2))
            voltage_rmse = np.std(voltages)
            voltage_rms_scaled = voltage_rms * voltage_scale
            voltage_rmse_scaled = voltage_rmse * voltage_scale

            self.ax_right.plot(
                times, voltages_scaled,
                color=plot_settings.color_voltage, linewidth=1.5, alpha=0.7, label='Voltage'
            )
            self.ax_right.set_ylim(bottom=0)
            self.ax_right.set_ylabel(
                f'Voltage ({voltage_unit})',
                fontsize=plot_settings.font_size_axis_label,
                fontweight='bold' if plot_settings.axis_label_bold else 'normal',
                fontfamily=font_family,
                color=plot_settings.color_voltage,
            )
            self.ax_right.tick_params(axis='y', labelcolor=plot_settings.color_voltage)
            self.ax_right.set_yscale('linear')

            if voltage_rmse < 0.25 * voltage_rms:
                self.ax_right.set_ylim(top=np.max(voltages_scaled) * 1.3)

            self.ax_right.text(
                0.02, 0.97,
                f'RMS = {voltage_rms_scaled:.2f} {voltage_unit}\n'
                f'RMSE = {voltage_rmse_scaled:.2f} {voltage_unit}',
                transform=self.ax_right.transAxes,
                fontsize=plot_settings.font_size_legend,
                fontfamily=font_family,
                verticalalignment='top',
                color=plot_settings.color_voltage,
                bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                          edgecolor=plot_settings.color_voltage, alpha=0.8),
            )

        elif b_channel == "resistance":
            self.ax_right = self.ax.twinx()
            self.ax_right.set_zorder(self.ax.get_zorder() - 1)
            self.ax.patch.set_visible(False)

            resistance_unit, resistance_scale = get_resistance_unit_and_scale(np.max(resistances))
            resistances_scaled = resistances * resistance_scale

            self.ax_right.plot(
                times, resistances_scaled,
                color=plot_settings.color_resistance, linewidth=1.5, alpha=0.7, label='Resistance'
            )
            self.ax_right.set_ylabel(
                f'Resistance ({resistance_unit})',
                fontsize=plot_settings.font_size_axis_label,
                fontweight='bold' if plot_settings.axis_label_bold else 'normal',
                fontfamily=font_family,
                color=plot_settings.color_resistance,
            )
            self.ax_right.tick_params(axis='y', labelcolor=plot_settings.color_resistance)

            for r in rounds:
                if r.res_info is not None:
                    pre_start = max(0, r.start_idx - 24)
                    pre_mid = int((pre_start + r.start_idx) / 2)
                    self.ax_right.plot(
                        times[pre_mid], r.res_info.r_max * resistance_scale,
                        'v', color=plot_settings.color_resistance,
                        markersize=6, alpha=0.8,
                    )
                    post_mid = int((r.peak_idx + min(r.peak_idx + 24, r.end_idx)) / 2)
                    self.ax_right.plot(
                        times[post_mid], r.res_info.r_min * resistance_scale,
                        '^', color=plot_settings.color_resistance,
                        markersize=6, alpha=0.8,
                    )

            self.ax_right.set_yscale(right_scale)

        rl = analysis_settings.rise_lower_percent
        ru = analysis_settings.rise_upper_percent
        fu = analysis_settings.fall_upper_percent
        fl = analysis_settings.fall_lower_percent

        legend_elements = [
            Line2D([0], [0], color=plot_settings.color_rise, linewidth=2, label='Rising edge'),
            Line2D([0], [0], color=plot_settings.color_fall, linewidth=2, label='Falling edge'),
            Line2D([0], [0], marker='*', color=plot_settings.color_peak, linestyle='None',
                   markersize=10, label='Peak'),
            Line2D([0], [0], marker='o', color=plot_settings.color_rise_marker, linestyle='None',
                   markersize=5, label=f'{rl:.0f}%(P.E.)'),
            Line2D([0], [0], marker='s', color=plot_settings.color_rise_marker, linestyle='None',
                   markersize=5, label=f'{ru:.0f}%(P.E.)'),
            Line2D([0], [0], marker='o', color=plot_settings.color_fall_marker, linestyle='None',
                   markersize=5, label=f'{fu:.0f}%(N.E.)'),
            Line2D([0], [0], marker='s', color=plot_settings.color_fall_marker, linestyle='None',
                   markersize=5, label=f'{fl:.0f}%(N.E.)'),
        ]

        if b_channel == "voltage":
            legend_elements.append(
                Line2D([0], [0], color=plot_settings.color_voltage, linewidth=1.5, label='Voltage')
            )
        elif b_channel == "resistance":
            legend_elements.append(
                Line2D([0], [0], color=plot_settings.color_resistance, linewidth=1.5, label='Resistance')
            )

        legend = self.ax.legend(
            handles=legend_elements, loc='upper right',
            fontsize=plot_settings.font_size_legend, prop={'family': font_family},
        )
        legend.set_zorder(10)

        self.ax.set_xlabel(
            'Time (s)',
            fontsize=plot_settings.font_size_axis_label,
            fontweight='bold' if plot_settings.axis_label_bold else 'normal',
            fontfamily=font_family,
        )
        self.ax.set_title(
            'Gas Sensor Experiment',
            fontsize=plot_settings.font_size_title,
            fontweight='bold' if plot_settings.title_bold else 'normal',
            fontfamily=font_family,
        )

        self.ax.tick_params(axis='both', labelsize=plot_settings.font_size_tick)
        for label in self.ax.get_xticklabels() + self.ax.get_yticklabels():
            label.set_fontfamily('Times New Roman')

        if self.ax_right:
            self.ax_right.tick_params(axis='y', labelsize=plot_settings.font_size_tick)
            for label in self.ax_right.get_yticklabels():
                label.set_fontfamily('Times New Roman')

        self.ax.grid(True, which='both', alpha=0.3)
        self.fig.tight_layout()
        self.draw()

    def get_figure(self) -> Figure:
        """返回 Figure 对象"""
        return self.fig


class MainWindow(QMainWindow):
    """主窗口"""

    def __init__(self):
        super().__init__()
        self.settings_manager = SettingsManager()
        self.lang = LanguageManager(language=self.settings_manager.get_settings().language)
        self.analysis_result: Optional[AnalysisResult] = None
        self.current_file_path: Optional[str] = None
        self._setup_ui()
        self._setup_menu()
        self._setup_status_bar()

    def _tr(self, key: str, **kwargs) -> str:
        return self.lang.get(key, **kwargs)

    def _setup_ui(self):
        """设置界面布局"""
        self.setWindowTitle(self._tr("window_title"))
        self.setMinimumSize(1400, 900)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(5, 5, 5, 5)
        main_layout.setSpacing(10)

        left_panel = QFrame()
        left_panel.setFixedWidth(200)
        left_panel.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #d0d7de;
                border-radius: 6px;
            }
        """)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(10, 15, 10, 10)

        # 数据通道选择组
        self.channel_group = QGroupBox(self._tr("group_channel_select"))
        channel_layout = QVBoxLayout(self.channel_group)

        self.channel_a_label = QLabel(self._tr("label_channel_a"))
        self.channel_a_label.setStyleSheet("font-weight: bold; color: #0969da;")
        channel_layout.addWidget(self.channel_a_label)

        self.channel_b_label = QLabel(self._tr("label_channel_b"))
        self.channel_b_label.setStyleSheet("font-weight: bold;")
        channel_layout.addWidget(self.channel_b_label)

        self.combo_b_channel = QComboBox()
        self.combo_b_channel.addItem(self._tr("opt_b_none"), "")
        self.combo_b_channel.addItem(self._tr("opt_b_resistance"), "resistance")
        self.combo_b_channel.addItem(self._tr("opt_b_voltage"), "voltage")
        saved_b = self.settings_manager.get_settings().b_channel
        idx = self.combo_b_channel.findData(saved_b)
        self.combo_b_channel.setCurrentIndex(max(idx, 0))
        self.combo_b_channel.currentIndexChanged.connect(self._on_b_channel_changed)
        channel_layout.addWidget(self.combo_b_channel)

        left_layout.addWidget(self.channel_group)

        # 坐标轴缩放选项组
        self.scale_group = QGroupBox(self._tr("group_axis_scale"))
        scale_layout = QVBoxLayout(self.scale_group)

        self.left_label = QLabel(self._tr("label_left_axis"))
        self.left_label.setStyleSheet("font-weight: bold; color: #0969da;")
        scale_layout.addWidget(self.left_label)

        self.combo_left_scale = QComboBox()
        self.combo_left_scale.addItems([self._tr("scale_linear"), self._tr("scale_log")])
        self.combo_left_scale.currentIndexChanged.connect(self._on_scale_changed)
        scale_layout.addWidget(self.combo_left_scale)

        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.HLine)
        separator.setStyleSheet("background-color: #d0d7de;")
        scale_layout.addWidget(separator)

        self.right_label = QLabel(self._tr("label_right_axis"))
        self.right_label.setStyleSheet("font-weight: bold; color: #6e7781;")
        scale_layout.addWidget(self.right_label)

        self.combo_right_scale = QComboBox()
        self.combo_right_scale.addItems([self._tr("scale_linear"), self._tr("scale_log")])
        self.combo_right_scale.currentIndexChanged.connect(self._on_scale_changed)
        scale_layout.addWidget(self.combo_right_scale)

        left_layout.addWidget(self.scale_group)
        left_layout.addStretch()
        main_layout.addWidget(left_panel)

        # 右侧内容区域
        right_content = QWidget()
        right_layout = QVBoxLayout(right_content)
        right_layout.setContentsMargins(0, 0, 0, 0)

        splitter = QSplitter(Qt.Vertical)

        self.plot_canvas = PlotCanvas()
        splitter.addWidget(self.plot_canvas)

        self.data_table = QTableWidget()
        self.data_table.setColumnCount(10)
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.data_table.setHorizontalScrollMode(QTableWidget.ScrollMode.ScrollPerPixel)
        self.data_table.setAlternatingRowColors(True)
        self.data_table.setEditTriggers(QTableWidget.NoEditTriggers)
        splitter.addWidget(self.data_table)

        splitter.setSizes([600, 200])
        right_layout.addWidget(splitter)

        button_layout = QHBoxLayout()
        button_layout.addStretch()

        self.btn_save_svg = QPushButton(self._tr("btn_save_svg"))
        self.btn_save_svg.clicked.connect(self._save_svg)
        button_layout.addWidget(self.btn_save_svg)

        self.btn_save_png = QPushButton(self._tr("btn_save_png"))
        self.btn_save_png.clicked.connect(self._save_png)
        button_layout.addWidget(self.btn_save_png)

        self.btn_copy_clipboard = QPushButton(self._tr("btn_copy"))
        self.btn_copy_clipboard.clicked.connect(self._copy_to_clipboard)
        button_layout.addWidget(self.btn_copy_clipboard)

        right_layout.addLayout(button_layout)
        main_layout.addWidget(right_content)

        self._sync_right_axis_state()

    def _setup_menu(self):
        """设置菜单栏"""
        menubar = self.menuBar()

        # 文件菜单
        file_menu = menubar.addMenu(self._tr("menu_file"))

        open_action = QAction(self._tr("menu_open"), self)
        open_action.setShortcut("Ctrl+O")
        open_action.triggered.connect(self._open_file)
        file_menu.addAction(open_action)

        file_menu.addSeparator()

        batch_action = QAction(self._tr("menu_batch"), self)
        batch_action.setShortcut("Ctrl+B")
        batch_action.triggered.connect(self._batch_process)
        file_menu.addAction(batch_action)

        file_menu.addSeparator()

        exit_action = QAction(self._tr("menu_exit"), self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # 设置菜单
        settings_menu = menubar.addMenu(self._tr("menu_settings"))

        system_prefs_action = QAction(self._tr("menu_system_prefs"), self)
        system_prefs_action.setShortcut("Ctrl+,")
        system_prefs_action.triggered.connect(self._show_system_settings)
        settings_menu.addAction(system_prefs_action)

        plot_prefs_action = QAction(self._tr("menu_plot_prefs"), self)
        plot_prefs_action.setShortcut("Ctrl+Shift+,")
        plot_prefs_action.triggered.connect(self._show_plot_settings)
        settings_menu.addAction(plot_prefs_action)

    def _setup_status_bar(self):
        """设置状态栏"""
        self.statusBar().showMessage(self._tr("status_ready"))

    def _apply_gui_scale(self, scale: float):
        """应用GUI缩放比例"""
        app: QApplication = QApplication.instance()  # type: ignore[assignment]
        font = QFont("Segoe UI", int(11 * scale))
        app.setFont(font)
        scaled_style = FLAT_STYLE.replace("font-size: 11px", f"font-size: {int(11 * scale)}px")
        app.setStyleSheet(scaled_style)

    def _b_channel_value(self) -> str:
        """获取当前B通道选择值"""
        return self.combo_b_channel.currentData()

    def _sync_right_axis_state(self):
        """根据B通道选择同步右侧Y轴状态"""
        b_channel = self._b_channel_value()
        enabled = b_channel != ""
        self.combo_right_scale.setEnabled(enabled)

        plot_colors = self.settings_manager.get_settings().plot
        if b_channel == "voltage":
            self.right_label.setStyleSheet(f"font-weight: bold; color: {plot_colors.color_voltage};")
        elif b_channel == "resistance":
            self.right_label.setStyleSheet(f"font-weight: bold; color: {plot_colors.color_resistance};")
        else:
            self.right_label.setStyleSheet("font-weight: bold; color: #6e7781;")

    def _on_b_channel_changed(self):
        """B通道选择改变时的处理"""
        b_channel = self._b_channel_value()
        self.settings_manager.update_b_channel(b_channel)
        self._sync_right_axis_state()
        if self.analysis_result and self.current_file_path:
            self._run_analysis(self.current_file_path)
        else:
            self._refresh_plot()
            self._update_table()

    def _on_scale_changed(self):
        """坐标轴缩放改变时的处理"""
        self._refresh_plot()

    def _refresh_plot(self):
        """刷新绘图"""
        if self.analysis_result:
            settings = self.settings_manager.get_settings()
            b_channel = self._b_channel_value()
            left_scale = 'log' if self.combo_left_scale.currentIndex() == 1 else 'linear'
            right_scale = 'log' if self.combo_right_scale.currentIndex() == 1 else 'linear'

            self.plot_canvas.plot_data(
                self.analysis_result, settings,
                b_channel=b_channel,
                left_scale=left_scale,
                right_scale=right_scale,
            )

    def _open_file(self):
        """打开文件对话框"""
        last_dir = self.settings_manager.get_settings().last_open_dir
        filepath, _ = QFileDialog.getOpenFileName(
            self, self._tr("dlg_open_title"), last_dir,
            self._tr("dlg_open_filter")
        )

        if filepath:
            self.settings_manager.update_last_open_dir(str(Path(filepath).parent))
            self.current_file_path = filepath
            self._run_analysis(filepath)

    def _run_analysis(self, filepath: str):
        """运行分析"""
        self.statusBar().showMessage(self._tr("status_analyzing", filename=Path(filepath).name))
        QApplication.processEvents()

        try:
            settings = self.settings_manager.get_settings()
            b_channel = self._b_channel_value()
            self.analysis_result = run_analysis(
                filepath,
                rise_lower_percent=settings.analysis.rise_lower_percent,
                rise_upper_percent=settings.analysis.rise_upper_percent,
                fall_upper_percent=settings.analysis.fall_upper_percent,
                fall_lower_percent=settings.analysis.fall_lower_percent,
                prominence=settings.analysis.prominence,
                distance=settings.analysis.distance,
                baseline_fraction=settings.analysis.baseline_fraction,
                peak_threshold=settings.analysis.peak_threshold,
                b_channel=b_channel,
            )

            self._refresh_plot()
            self._update_table()

            self.statusBar().showMessage(
                self._tr("status_analyzed", filename=Path(filepath).name, n=len(self.analysis_result.rounds))
            )
        except Exception as e:
            QMessageBox.critical(self, self._tr("msg_error"), self._tr("msg_analysis_failed", error=str(e)))
            self.statusBar().showMessage(self._tr("status_failed"))

    def _update_table(self):
        """更新数据表格（每列使用独立单位以避免精度显示丢失）"""
        if not self.analysis_result:
            return

        rounds = self.analysis_result.rounds
        analysis_settings = self.settings_manager.get_settings().analysis
        b_channel = self._b_channel_value()

        rl = analysis_settings.rise_lower_percent
        ru = analysis_settings.rise_upper_percent
        fu = analysis_settings.fall_upper_percent
        fl = analysis_settings.fall_lower_percent

        # --- 收集各列数据以确定独立单位 ---
        peak_vals = [r.peak_current for r in rounds]
        idle_vals = [r.rise.min_val for r in rounds if r.rise]
        rise_upper_vals = [r.rise.i_at_upper for r in rounds if r.rise]
        rise_lower_vals = [r.rise.i_at_lower for r in rounds if r.rise]
        fall_upper_vals = [r.fall.i_at_upper for r in rounds if r.fall]
        fall_lower_vals = [r.fall.i_at_lower for r in rounds if r.fall]

        peak_unit, peak_scale = _column_unit_scale(peak_vals, get_current_unit_and_scale)
        idle_unit, idle_scale = _column_unit_scale(idle_vals, get_current_unit_and_scale)
        ru_unit, ru_scale = _column_unit_scale(rise_upper_vals, get_current_unit_and_scale)
        rl_unit, rl_scale = _column_unit_scale(rise_lower_vals, get_current_unit_and_scale)
        fu_unit, fu_scale = _column_unit_scale(fall_upper_vals, get_current_unit_and_scale)
        fl_unit, fl_scale = _column_unit_scale(fall_lower_vals, get_current_unit_and_scale)

        headers = [
            self._tr("col_round"),
            f'{self._tr("col_peak_i")} ({peak_unit})',
            f'{self._tr("col_idle_i")} ({idle_unit})',
            self._tr("col_response"),
            self._tr("col_rise_time"),
            self._tr("col_fall_time"),
            f'{self._tr("col_rise_i_upper", percent=f"{ru:.0f}")} ({ru_unit})',
            f'{self._tr("col_rise_i_lower", percent=f"{rl:.0f}")} ({rl_unit})',
            f'{self._tr("col_fall_i_upper", percent=f"{fu:.0f}")} ({fu_unit})',
            f'{self._tr("col_fall_i_lower", percent=f"{fl:.0f}")} ({fl_unit})',
        ]

        # 通道B电阻列：独立单位
        rmax_unit = rmin_unit = ""
        rmax_scale = rmin_scale = 1.0
        if b_channel == "resistance":
            rmax_vals = [r.res_info.r_max for r in rounds if r.res_info]
            rmin_vals = [r.res_info.r_min for r in rounds if r.res_info]
            rmax_unit, rmax_scale = _column_unit_scale(rmax_vals, get_resistance_unit_and_scale)
            rmin_unit, rmin_scale = _column_unit_scale(rmin_vals, get_resistance_unit_and_scale)
            headers.append(f'{self._tr("col_r_max")} ({rmax_unit})')
            headers.append(f'{self._tr("col_r_min")} ({rmin_unit})')
            headers.append(self._tr("col_r_ratio"))

        self.data_table.setColumnCount(len(headers))
        self.data_table.setHorizontalHeaderLabels(headers)
        self.data_table.setRowCount(len(rounds))

        for i, r in enumerate(rounds):
            self.data_table.setItem(i, 0, QTableWidgetItem(str(r.round_num)))
            self.data_table.setItem(i, 1, QTableWidgetItem(f"{r.peak_current * peak_scale:.4f}"))

            idle_str = f"{r.rise.min_val * idle_scale:.4f}" if r.rise else "N/A"
            self.data_table.setItem(i, 2, QTableWidgetItem(idle_str))

            response_str = f"{r.rise.ratio:.2f}" if r.rise else "N/A"
            self.data_table.setItem(i, 3, QTableWidgetItem(response_str))

            rise_time_str = f"{r.rise.transition_time:.4f}" if r.rise else "N/A"
            fall_time_str = f"{r.fall.transition_time:.4f}" if r.fall else "N/A"
            self.data_table.setItem(i, 4, QTableWidgetItem(rise_time_str))
            self.data_table.setItem(i, 5, QTableWidgetItem(fall_time_str))

            if r.rise:
                self.data_table.setItem(i, 6, QTableWidgetItem(f"{r.rise.i_at_upper * ru_scale:.4f}"))
                self.data_table.setItem(i, 7, QTableWidgetItem(f"{r.rise.i_at_lower * rl_scale:.4f}"))
            else:
                self.data_table.setItem(i, 6, QTableWidgetItem("N/A"))
                self.data_table.setItem(i, 7, QTableWidgetItem("N/A"))

            if r.fall:
                self.data_table.setItem(i, 8, QTableWidgetItem(f"{r.fall.i_at_upper * fu_scale:.4f}"))
                self.data_table.setItem(i, 9, QTableWidgetItem(f"{r.fall.i_at_lower * fl_scale:.4f}"))
            else:
                self.data_table.setItem(i, 8, QTableWidgetItem("N/A"))
                self.data_table.setItem(i, 9, QTableWidgetItem("N/A"))

            if b_channel == "resistance":
                col_offset = 10
                if r.res_info:
                    self.data_table.setItem(i, col_offset,
                        QTableWidgetItem(f"{r.res_info.r_max * rmax_scale:.2f}"))
                    self.data_table.setItem(i, col_offset + 1,
                        QTableWidgetItem(f"{r.res_info.r_min * rmin_scale:.2f}"))
                    self.data_table.setItem(i, col_offset + 2,
                        QTableWidgetItem(f"{r.res_info.r_ratio:.2f}"))
                else:
                    self.data_table.setItem(i, col_offset, QTableWidgetItem("N/A"))
                    self.data_table.setItem(i, col_offset + 1, QTableWidgetItem("N/A"))
                    self.data_table.setItem(i, col_offset + 2, QTableWidgetItem("N/A"))

    # ── 批量处理 ──────────────────────────────────────────────────

    def _batch_process(self):
        """批量处理入口"""
        last_dir = self.settings_manager.get_settings().last_open_dir
        filepaths, _ = QFileDialog.getOpenFileNames(
            self, self._tr("dlg_batch_title"), last_dir,
            self._tr("dlg_open_filter")
        )
        if not filepaths:
            return

        # 步骤1：确认分析设置
        settings = self.settings_manager.get_settings()
        confirm_dlg = BatchConfirmDialog(settings, len(filepaths), self.lang, self)
        if confirm_dlg.exec() != QDialog.Accepted:
            return

        # 步骤2：输出设置
        output_dlg = BatchOutputDialog(self.lang, self)
        if output_dlg.exec() != QDialog.Accepted:
            return

        excel_name = output_dlg.edit_excel_name.text().strip()
        output_dir = output_dlg.edit_output_dir.text().strip()
        save_svg = output_dlg.get_svg_enabled()
        save_png = output_dlg.get_png_enabled()

        if not output_dir:
            QMessageBox.warning(self, self._tr("msg_warning"), self._tr("msg_no_output_dir"))
            return

        out_path = Path(output_dir)
        if not out_path.exists():
            try:
                out_path.mkdir(parents=True)
            except OSError as e:
                QMessageBox.critical(self, self._tr("msg_error"),
                                     self._tr("msg_dir_create_failed", error=str(e)))
                return

        # 步骤3：批量处理
        n_total = len(filepaths)
        progress = QProgressDialog(
            self._tr("status_batch_processing"), self._tr("btn_cancel"), 0, n_total, self
        )
        progress.setWindowTitle(self._tr("dlg_batch_title"))
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)

        b_channel = self._b_channel_value()
        analysis_settings = settings.analysis
        all_results: List[AnalysisResult] = []

        for idx, fp in enumerate(filepaths):
            if progress.wasCanceled():
                break

            progress.setValue(idx)
            progress.setLabelText(
                self._tr("status_batch_file", filename=Path(fp).name, current=idx + 1, total=n_total)
            )
            QApplication.processEvents()

            try:
                result = run_analysis(
                    fp,
                    rise_lower_percent=analysis_settings.rise_lower_percent,
                    rise_upper_percent=analysis_settings.rise_upper_percent,
                    fall_upper_percent=analysis_settings.fall_upper_percent,
                    fall_lower_percent=analysis_settings.fall_lower_percent,
                    prominence=analysis_settings.prominence,
                    distance=analysis_settings.distance,
                    baseline_fraction=analysis_settings.baseline_fraction,
                    peak_threshold=analysis_settings.peak_threshold,
                    b_channel=b_channel,
                )
                all_results.append(result)

                # 保存图像
                self._save_batch_plots(result, out_path, save_svg, save_png)

            except Exception as e:
                QMessageBox.warning(self, self._tr("msg_error"),
                                    self._tr("msg_batch_file_failed", filename=Path(fp).name, error=str(e)))

        progress.setValue(n_total)

        # 步骤4：导出汇总Excel
        if all_results:
            excel_path = out_path / excel_name
            try:
                self._export_batch_excel(all_results, excel_path, analysis_settings, b_channel)
                self.statusBar().showMessage(
                    self._tr("status_batch_done", n=len(all_results), path=str(excel_path))
                )
                QMessageBox.information(self, self._tr("msg_batch_complete"),
                                        self._tr("msg_batch_complete_detail",
                                                 n=len(all_results), total=n_total,
                                                 path=str(excel_path)))
            except Exception as e:
                QMessageBox.critical(self, self._tr("msg_error"),
                                     self._tr("msg_export_failed", error=str(e)))
        else:
            self.statusBar().showMessage(self._tr("status_batch_failed"))

    def _save_batch_plots(self, result: AnalysisResult, out_path: Path,
                          save_svg: bool, save_png: bool):
        """为单个文件的批量处理保存图像"""
        base_name = Path(result.file_path).stem

        # 创建一个临时 canvas 用于渲染
        canvas = PlotCanvas()
        settings = self.settings_manager.get_settings()
        b_channel = self._b_channel_value()

        # 使用全局一致的缩放设置
        left_scale = 'log' if self.combo_left_scale.currentIndex() == 1 else 'linear'
        right_scale = 'log' if self.combo_right_scale.currentIndex() == 1 else 'linear'

        canvas.plot_data(result, settings, b_channel=b_channel,
                         left_scale=left_scale, right_scale=right_scale)

        fig = canvas.get_figure()

        if save_svg:
            svg_path = out_path / f"{base_name}_plot.svg"
            fig.savefig(svg_path, format='svg', bbox_inches='tight')
        if save_png:
            png_path = out_path / f"{base_name}_plot.png"
            fig.savefig(png_path, format='png', dpi=150, bbox_inches='tight')

    def _export_batch_excel(self, results: List[AnalysisResult], excel_path: Path,
                            analysis_settings: AnalysisSettings, b_channel: str):
        """将批量分析结果导出为格式化的 Excel 汇总表"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Batch Results"

        # 样式定义
        header_fill = PatternFill(start_color="0969da", end_color="0969da", fill_type="solid")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        file_fill = PatternFill(start_color="f6f8fa", end_color="f6f8fa", fill_type="solid")
        file_font = Font(name="Segoe UI", size=11, bold=True, color="24292f")
        alt_fill = PatternFill(start_color="ddf4ff", end_color="ddf4ff", fill_type="solid")
        normal_font = Font(name="Segoe UI", size=9, color="24292f")
        thin_border = Border(
            left=Side(style='thin', color='d0d7de'),
            right=Side(style='thin', color='d0d7de'),
            top=Side(style='thin', color='d0d7de'),
            bottom=Side(style='thin', color='d0d7de'),
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        rl = analysis_settings.rise_lower_percent
        ru = analysis_settings.rise_upper_percent
        fu = analysis_settings.fall_upper_percent
        fl = analysis_settings.fall_lower_percent

        current_row = 1

        for result_idx, result in enumerate(results):
            rounds = result.rounds
            if not rounds:
                continue

            # --- 收集数据确定每列单位 ---
            peak_vals = [r.peak_current for r in rounds]
            idle_vals = [r.rise.min_val for r in rounds if r.rise]
            rise_upper_vals = [r.rise.i_at_upper for r in rounds if r.rise]
            rise_lower_vals = [r.rise.i_at_lower for r in rounds if r.rise]
            fall_upper_vals = [r.fall.i_at_upper for r in rounds if r.fall]
            fall_lower_vals = [r.fall.i_at_lower for r in rounds if r.fall]

            peak_unit, peak_scale = _column_unit_scale(peak_vals, get_current_unit_and_scale)
            idle_unit, idle_scale = _column_unit_scale(idle_vals, get_current_unit_and_scale)
            ru_unit, ru_scale = _column_unit_scale(rise_upper_vals, get_current_unit_and_scale)
            rl_unit, rl_scale = _column_unit_scale(rise_lower_vals, get_current_unit_and_scale)
            fu_unit, fu_scale = _column_unit_scale(fall_upper_vals, get_current_unit_and_scale)
            fl_unit, fl_scale = _column_unit_scale(fall_lower_vals, get_current_unit_and_scale)

            rmax_unit = rmin_unit = ""
            rmax_scale = rmin_scale = 1.0
            has_resistance = (b_channel == "resistance" and
                              any(r.res_info is not None for r in rounds))
            if has_resistance:
                rmax_vals = [r.res_info.r_max for r in rounds if r.res_info]
                rmin_vals = [r.res_info.r_min for r in rounds if r.res_info]
                rmax_unit, rmax_scale = _column_unit_scale(rmax_vals, get_resistance_unit_and_scale)
                rmin_unit, rmin_scale = _column_unit_scale(rmin_vals, get_resistance_unit_and_scale)

            # 构建表头
            headers = [
                "Round",
                f"Peak I ({peak_unit})",
                f"Idle I ({idle_unit})",
                "Response",
                "Rise Time (s)",
                "Fall Time (s)",
                f"I_{ru:.0f}% (Rise) ({ru_unit})",
                f"I_{rl:.0f}% (Rise) ({rl_unit})",
                f"I_{fu:.0f}% (Fall) ({fu_unit})",
                f"I_{fl:.0f}% (Fall) ({fl_unit})",
            ]
            if has_resistance:
                headers.append(f"Rmax ({rmax_unit})")
                headers.append(f"Rmin ({rmin_unit})")
                headers.append("Rmax/Rmin")

            n_cols = len(headers)

            # --- 文件名行（合并单元格） ---
            if result_idx > 0:
                current_row += 1  # 文件间空行
            file_name = Path(result.file_path).name
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=n_cols)
            cell = ws.cell(row=current_row, column=1, value=file_name)
            cell.fill = file_fill
            cell.font = file_font
            cell.alignment = left_align
            for c in range(1, n_cols + 1):
                ws.cell(row=current_row, column=c).fill = file_fill
                ws.cell(row=current_row, column=c).border = thin_border
            current_row += 1

            # --- 表头行 ---
            for col_idx, hdr in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=hdr)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1

            # --- 数据行 ---
            for i, r in enumerate(rounds):
                row_data = [
                    r.round_num,
                    r.peak_current * peak_scale,
                    r.rise.min_val * idle_scale if r.rise else None,
                    r.rise.ratio if r.rise else None,
                    r.rise.transition_time if r.rise else None,
                    r.fall.transition_time if r.fall else None,
                    r.rise.i_at_upper * ru_scale if r.rise else None,
                    r.rise.i_at_lower * rl_scale if r.rise else None,
                    r.fall.i_at_upper * fu_scale if r.fall else None,
                    r.fall.i_at_lower * fl_scale if r.fall else None,
                ]
                if has_resistance:
                    if r.res_info:
                        row_data.append(r.res_info.r_max * rmax_scale)
                        row_data.append(r.res_info.r_min * rmin_scale)
                        row_data.append(r.res_info.r_ratio)
                    else:
                        row_data.extend([None, None, None])

                row_fill = alt_fill if i % 2 == 0 else None
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    if val is not None:
                        cell.value = val
                    else:
                        cell.value = "N/A"
                    cell.font = normal_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    if row_fill:
                        cell.fill = row_fill

                # 格式化数值单元格
                if row_data[1] is not None:
                    ws.cell(row=current_row, column=2).number_format = '0.0000'
                if row_data[2] is not None:
                    ws.cell(row=current_row, column=3).number_format = '0.0000'
                if row_data[3] is not None:
                    ws.cell(row=current_row, column=4).number_format = '0.00'
                if row_data[4] is not None:
                    ws.cell(row=current_row, column=5).number_format = '0.0000'
                if row_data[5] is not None:
                    ws.cell(row=current_row, column=6).number_format = '0.0000'
                if row_data[6] is not None:
                    ws.cell(row=current_row, column=7).number_format = '0.0000'
                if row_data[7] is not None:
                    ws.cell(row=current_row, column=8).number_format = '0.0000'
                if row_data[8] is not None:
                    ws.cell(row=current_row, column=9).number_format = '0.0000'
                if row_data[9] is not None:
                    ws.cell(row=current_row, column=10).number_format = '0.0000'
                if has_resistance:
                    if row_data[10] is not None:
                        ws.cell(row=current_row, column=11).number_format = '0.00'
                    if row_data[11] is not None:
                        ws.cell(row=current_row, column=12).number_format = '0.00'
                    if row_data[12] is not None:
                        ws.cell(row=current_row, column=13).number_format = '0.00'

                current_row += 1

        # 调整列宽
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        wb.save(excel_path)

    # ── 设置对话框 ────────────────────────────────────────────────

    def _show_system_settings(self):
        """显示系统首选项对话框"""
        dialog = SystemSettingsDialog(self.settings_manager.get_settings(), self.lang, self)
        if dialog.exec() == QDialog.Accepted:
            new_settings = dialog.get_settings()
            analysis_changed = dialog.get_analysis_changed()

            self.settings_manager.update_analysis(**new_settings['analysis'])
            self.settings_manager.update_gui_scale(new_settings['gui_scale'])
            self.settings_manager.update_language(new_settings['language'])

            old_lang = self.lang.language
            new_lang = new_settings['language']
            if new_lang != old_lang:
                self.lang.set_language(new_lang)
                self._retranslate_ui()

            self.settings_manager.load()

            scale = new_settings['gui_scale']
            self._apply_gui_scale(scale)

            if self.analysis_result:
                if analysis_changed and self.current_file_path:
                    self._run_analysis(self.current_file_path)
                else:
                    self._refresh_plot()
                    self._update_table()

    def _show_plot_settings(self):
        """显示绘图首选项对话框"""
        dialog = PlotSettingsDialog(self.settings_manager.get_settings(), self.lang, self)
        if dialog.exec() == QDialog.Accepted:
            new_settings = dialog.get_settings()
            self.settings_manager.update_plot(**new_settings['plot'])
            self.settings_manager.load()

            if self.analysis_result:
                self._refresh_plot()

    # ── 翻译 ──────────────────────────────────────────────────────

    def _retranslate_ui(self):
        """重新翻译所有UI文本"""
        self.setWindowTitle(self._tr("window_title"))
        self.statusBar().showMessage(self._tr("status_ready"))

        self.channel_group.setTitle(self._tr("group_channel_select"))
        self.channel_a_label.setText(self._tr("label_channel_a"))
        self.channel_b_label.setText(self._tr("label_channel_b"))
        self.scale_group.setTitle(self._tr("group_axis_scale"))
        self.left_label.setText(self._tr("label_left_axis"))
        self.right_label.setText(self._tr("label_right_axis"))

        self.combo_b_channel.setItemText(0, self._tr("opt_b_none"))
        self.combo_b_channel.setItemText(1, self._tr("opt_b_resistance"))
        self.combo_b_channel.setItemText(2, self._tr("opt_b_voltage"))

        self.btn_save_svg.setText(self._tr("btn_save_svg"))
        self.btn_save_png.setText(self._tr("btn_save_png"))
        self.btn_copy_clipboard.setText(self._tr("btn_copy"))

        self._update_table()

        self.menuBar().clear()
        self._setup_menu()

    # ── 保存 / 复制 ───────────────────────────────────────────────

    def _save_svg(self):
        """保存为 SVG"""
        if not self.analysis_result:
            QMessageBox.warning(self, self._tr("msg_warning"), self._tr("msg_no_data_save"))
            return

        filepath, _ = QFileDialog.getSaveFileName(
            self, self._tr("dlg_save_svg_title"), "", self._tr("dlg_save_svg_filter")
        )
        if filepath:
            self.plot_canvas.get_figure().savefig(filepath, format='svg', bbox_inches='tight')
            self.statusBar().showMessage(self._tr("status_saved", filepath=filepath))

    def _save_png(self):
        """保存为 PNG"""
        if not self.analysis_result:
            QMessageBox.warning(self, self._tr("msg_warning"), self._tr("msg_no_data_save"))
            return

        filepath, _ = QFileDialog.getSaveFileName(
            self, self._tr("dlg_save_png_title"), "", self._tr("dlg_save_png_filter")
        )
        if filepath:
            self.plot_canvas.get_figure().savefig(filepath, format='png', dpi=150, bbox_inches='tight')
            self.statusBar().showMessage(self._tr("status_saved", filepath=filepath))

    def _copy_to_clipboard(self):
        """复制图片到剪贴板（PNG格式）"""
        if not self.analysis_result:
            QMessageBox.warning(self, self._tr("msg_warning"), self._tr("msg_no_data_copy"))
            return

        from PySide6.QtGui import QImage
        from io import BytesIO

        buf = BytesIO()
        self.plot_canvas.get_figure().savefig(buf, format='png', dpi=150, bbox_inches='tight')
        buf.seek(0)

        image = QImage()
        image.loadFromData(buf.getvalue())

        clipboard = QApplication.clipboard()
        clipboard.setImage(image)

        self.statusBar().showMessage(self._tr("status_copied"))


def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Gas Sensor Analyzer")
    app.setOrganizationName("GasSensor")

    app.setStyle("Fusion")

    sm = SettingsManager()
    gui_scale = sm.get_settings().gui_scale
    scaled_style = FLAT_STYLE.replace("font-size: 11px", f"font-size: {int(11 * gui_scale)}px")
    app.setStyleSheet(scaled_style)

    font = QFont("Segoe UI", int(11 * gui_scale))
    app.setFont(font)

    window = MainWindow()
    window.show()

    sys.exit(app.exec())


if __name__ == '__main__':
    main()
