"""
核心分析模块。
包含数据加载、平滑、轮次切分、边沿分析和B通道分析逻辑。
采用管道式架构：加载 → 平滑 → 分割 → 边沿提取 → B通道分析。
"""

import numpy as np
import openpyxl
from scipy.signal import savgol_filter, find_peaks
from dataclasses import dataclass
from typing import List, Optional, Tuple


@dataclass
class EdgeInfo:
    """边沿信息（通道A：电流）"""
    max_val: float
    min_val: float
    ratio: float
    t_lower: float          # 下限百分比对应的时间
    t_upper: float          # 上限百分比对应的时间
    transition_time: float
    i_at_lower: float       # t_lower 时刻的电流值
    i_at_upper: float       # t_upper 时刻的电流值


@dataclass
class OvershootInfo:
    """过冲峰分析信息"""
    rapid_descent_start_idx: int       # 快速下降起始点索引
    rapid_descent_start_time: float    # 快速下降起始时间
    i_90_stable: float                 # 过冲下降结束的稳定电流值（I_90_Stable）
    slow_descent_time: float           # 慢下降时间（从峰值到I_90_Stable交叉点）
    overshoot_descent_time: float      # 从上升起点到过冲下降达I_90_Stable的时间


@dataclass
class ResistanceRoundInfo:
    """单轮电阻分析结果（通道B：电阻）"""
    r_max: float            # 上升沿前一区间电阻稳健最大值
    r_min: float            # 上升沿后区间电阻稳健最小值
    r_ratio: float          # r_max / r_min


@dataclass
class RoundResult:
    """单轮分析结果"""
    round_num: int
    peak_time: float
    peak_current: float
    rise: Optional[EdgeInfo]
    fall: Optional[EdgeInfo]
    start_idx: int = 0
    peak_idx: int = 0
    end_idx: int = 0
    res_info: Optional[ResistanceRoundInfo] = None
    is_overshoot: bool = False
    overshoot_info: Optional[OvershootInfo] = None


@dataclass
class AnalysisResult:
    """完整分析结果"""
    times: np.ndarray
    voltages: np.ndarray
    currents: np.ndarray
    resistances: np.ndarray
    currents_smooth: np.ndarray
    rounds: List[RoundResult]
    file_path: str


# ── Stage 0: 数据加载 ──────────────────────────────────────────────

def load_data(filepath: str) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """读取 xlsx 文件，返回 (时间, 电压, 电流, 电阻) 数组。"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    times, voltages, currents, resistances = [], [], [], []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        times.append(row[1])
        voltages.append(row[2])
        currents.append(row[3])
        resistances.append(row[4])

    return (
        np.array(times, dtype=float),
        np.array(voltages, dtype=float),
        np.array(currents, dtype=float),
        np.array(resistances, dtype=float),
    )


# ── Stage 1: 信号平滑 ──────────────────────────────────────────────

def smooth_current(currents: np.ndarray, window: int = 21, poly: int = 3) -> np.ndarray:
    """Savitzky-Golay 平滑，消除单点毛刺。"""
    window = min(window, len(currents))
    if window % 2 == 0:
        window -= 1
    smoothed = savgol_filter(currents, window_length=window, polyorder=poly)
    return np.clip(smoothed, 0, None)


def smooth_signal(data: np.ndarray, window: int = 21, poly: int = 3) -> np.ndarray:
    """通用 Savitzky-Golay 平滑（不截断负值），用于电阻等非电流信号。"""
    window = min(window, len(data))
    if window % 2 == 0:
        window -= 1
    return savgol_filter(data, window_length=window, polyorder=poly)


# ── Stage 2: 实验阶段分割 ──────────────────────────────────────────

def find_round_boundaries(
    times: np.ndarray,
    currents_smooth: np.ndarray,
    prominence: float = 0.3,
    distance: int = 150,
    baseline_fraction: float = 0.01,
    peak_threshold: float = 1e-7,
) -> List[Tuple[int, int, int]]:
    """
    识别每一轮实验的 (start_idx, peak_idx, end_idx)。
    基于对数空间峰值检测，向左/右扫描至基线阈值。
    """
    log_current = np.log10(np.clip(currents_smooth, 1e-20, None))

    peaks, _ = find_peaks(log_current, prominence=prominence, distance=distance)

    valid_peaks = [p for p in peaks if currents_smooth[p] > peak_threshold]

    if len(valid_peaks) == 0:
        return []

    rounds = []

    for i, pk in enumerate(valid_peaks):
        peak_val = currents_smooth[pk]
        threshold = peak_val * baseline_fraction

        start = 0
        for j in range(pk, -1, -1):
            if currents_smooth[j] <= threshold:
                start = j
                break

        right_limit = len(times) - 1
        if i + 1 < len(valid_peaks):
            right_limit = valid_peaks[i + 1]

        end = right_limit
        for j in range(pk, min(right_limit + 1, len(times))):
            if currents_smooth[j] <= threshold:
                end = j
                break

        rounds.append((start, pk, end))

    return rounds


# ── Stage 3: 边沿提取 ──────────────────────────────────────────────

def interpolate_crossing_time(
    times: np.ndarray,
    values: np.ndarray,
    idx_before: int,
    idx_after: int,
    threshold: float,
) -> float:
    """线性插值找到精确的交叉时间。"""
    if idx_after == idx_before:
        return times[idx_before]
    t0, t1 = times[idx_before], times[idx_after]
    v0, v1 = values[idx_before], values[idx_after]
    if v1 == v0:
        return (t0 + t1) / 2.0
    frac = (threshold - v0) / (v1 - v0)
    frac = np.clip(frac, 0.0, 1.0)
    return t0 + frac * (t1 - t0)


def _find_crossing(
    seg_t: np.ndarray,
    seg_i: np.ndarray,
    threshold: float,
    direction: str,
) -> Optional[float]:
    """在线段中找到第一次穿过 threshold 的精确时刻。"""
    for k in range(len(seg_i) - 1):
        if direction == 'up':
            if seg_i[k] <= threshold <= seg_i[k + 1]:
                return interpolate_crossing_time(seg_t, seg_i, k, k + 1, threshold)
        else:
            if seg_i[k] >= threshold >= seg_i[k + 1]:
                return interpolate_crossing_time(seg_t, seg_i, k, k + 1, threshold)
    return None


def _trimmed_mean(data: np.ndarray, trim_fraction: float = 0.25) -> float:
    """去除最高和最低各 trim_fraction 比例后取平均值。"""
    n = len(data)
    if n == 0:
        return 0.0
    k = int(n * trim_fraction)
    if k * 2 >= n:
        return float(np.mean(data))
    sorted_data = np.sort(data)
    trimmed = sorted_data[k:n - k]
    return float(np.mean(trimmed))


def _trimmed_max(data: np.ndarray, n_points: int = 5) -> float:
    """取最大的 n_points 个值的平均值（稳健最大值）。"""
    if len(data) == 0:
        return 0.0
    k = min(n_points, len(data))
    return float(np.mean(np.sort(data)[-k:]))


def _trimmed_min(data: np.ndarray, n_points: int = 5) -> float:
    """取最小的 n_points 个值的平均值（稳健最小值）。"""
    if len(data) == 0:
        return 0.0
    k = min(n_points, len(data))
    return float(np.mean(np.sort(data)[:k]))


def analyze_edge(
    times: np.ndarray,
    currents_raw: np.ndarray,
    currents_smooth: np.ndarray,
    start_idx: int,
    end_idx: int,
    is_rising: bool,
    lower_percent: float = 10.0,
    upper_percent: float = 90.0,
) -> Optional[EdgeInfo]:
    """
    分析一段上升沿或下降沿（通道A：电流）。
    使用平滑数据确定交叉点和 min/max。
    上升沿的 I_min 取 start_idx 之前 24 点的截断均值。
    """
    seg_t = times[start_idx:end_idx + 1]
    seg_smooth = currents_smooth[start_idx:end_idx + 1]

    if len(seg_smooth) < 2:
        return None

    max_val = float(np.max(seg_smooth))

    if is_rising:
        baseline_start = max(0, start_idx - 24)
        baseline_window = currents_smooth[baseline_start:start_idx]
        if len(baseline_window) >= 4:
            min_val = _trimmed_mean(baseline_window, trim_fraction=0.25)
        elif len(baseline_window) > 0:
            min_val = float(np.mean(baseline_window))
        else:
            min_val = float(np.min(seg_smooth))
    else:
        min_val = float(np.min(seg_smooth))

    if max_val <= min_val:
        return None

    ratio = max_val / min_val if min_val > 0 else float('inf')

    level_lower = min_val + (lower_percent / 100.0) * (max_val - min_val)
    level_upper = min_val + (upper_percent / 100.0) * (max_val - min_val)

    if is_rising:
        t_lower = _find_crossing(seg_t, seg_smooth, level_lower, direction='up')
        t_upper = _find_crossing(seg_t, seg_smooth, level_upper, direction='up')
    else:
        t_upper = _find_crossing(seg_t, seg_smooth, level_upper, direction='down')
        t_lower = _find_crossing(seg_t, seg_smooth, level_lower, direction='down')

    if t_lower is None or t_upper is None:
        return None

    transition_time = abs(t_upper - t_lower)

    idx_lower = int(np.argmin(np.abs(times - t_lower)))
    idx_upper = int(np.argmin(np.abs(times - t_upper)))

    return EdgeInfo(
        max_val=max_val,
        min_val=min_val,
        ratio=ratio,
        t_lower=t_lower,
        t_upper=t_upper,
        transition_time=transition_time,
        i_at_lower=float(currents_smooth[idx_lower]),
        i_at_upper=float(currents_smooth[idx_upper]),
    )


# ── Stage 3a: 过冲峰检测 ──────────────────────────────────────────

def detect_peak_type(
    currents_smooth: np.ndarray,
    peak_idx: int,
    end_idx: int,
    start_idx: int = 0,
) -> bool:
    """
    判断单轮峰值类型。
    返回 True 表示过冲峰，False 表示饱和峰。

    原理（拟人化分析）：
    - 过冲峰：峰值后立即下降，下降先快后慢趋近稳态，气体结束时再快速下降。
      下降速率（负导数）在曲线上呈现两个"驼峰"（两个高下降斜率阶段）。
    - 饱和峰：峰值后电流基本维持（平台），仅在气体通入结束时快速下降。
      下降速率只有一个"驼峰"。

    判据（双阶段下降检测）：
    1. 计算下降段的负导数，找显著峰值
    2. 若前40%和后30%各有一个峰、且中间有明显下降 → 过冲峰
    3. 否则计算中点累积下降比：>0.3 → 过冲峰（下降分散），≤0.3 → 饱和峰（下降集中在末尾）
    """
    seg = currents_smooth[peak_idx:end_idx + 1]
    n = len(seg)
    if n < 10:
        return False

    # 负导数（下降速率）
    neg_deriv = -np.gradient(seg)
    max_d = np.max(neg_deriv)
    if max_d <= 0:
        return False

    # 找显著导数峰（高度 > max*5%，间距 > 15点）
    peaks_idx, _ = find_peaks(neg_deriv, height=max_d * 0.05)
    if len(peaks_idx) > 1:
        filtered = [peaks_idx[0]]
        for p in peaks_idx[1:]:
            if p - filtered[-1] >= 15:
                filtered.append(p)
        peaks_idx = np.array(filtered)

    # 双阶段检测：前40%有峰 + 后30%有峰 + 中间有明显下降
    two_phase = False
    if len(peaks_idx) >= 2:
        first, last = peaks_idx[0], peaks_idx[-1]
        if first < n * 0.4 and last > n * 0.7:
            valley = np.min(neg_deriv[first:last])
            if valley < neg_deriv[first] * 0.5 and valley < neg_deriv[last] * 0.5:
                two_phase = True

    if two_phase:
        return True

    # 降级判断：中点累积下降比
    diffs = np.diff(seg)
    cum = np.cumsum(np.maximum(0, -diffs))
    total = cum[-1] if len(cum) > 0 and cum[-1] > 0 else 1
    mid_ratio = cum[len(cum) // 2] / total if len(cum) > 0 else 0

    return mid_ratio > 0.3 and len(peaks_idx) >= 2


def find_rapid_descent_start(
    currents_smooth: np.ndarray,
    peak_idx: int,
    end_idx: int,
    remaining_threshold: float = 0.50,
) -> int:
    """
    在下降段 [peak_idx, end_idx] 中找到快速下降起始点。

    算法：
    1. 找到下降速率（负导数）最大的位置（快速下降的核心区域）
    2. 从该位置向后扫描，找累积下降已超过(1-threshold)的位置
       即剩余下降量不足总下降量的 threshold（默认50%）
    3. 该位置即为慢下降→快下降的转换点

    原理：过冲峰的慢下降阶段累积少量下降，快下降阶段累积大量下降。
    找到"大部分下降尚未发生"的最后时刻，即为快下降的起点。
    """
    seg = currents_smooth[peak_idx:end_idx + 1]
    n = len(seg)
    if n < 5:
        return peak_idx

    # 累积下降
    diffs = np.diff(seg)
    descent = np.maximum(0, -diffs)
    cum_descent = np.cumsum(descent)
    total_descent = cum_descent[-1] if cum_descent[-1] > 0 else 1

    # 找下降速率最大的位置
    max_deriv_idx = int(np.argmax(-np.gradient(seg)))

    # 从 max_deriv_idx 向后扫描，找 remaining < threshold
    for j in range(max_deriv_idx, -1, -1):
        remaining = 1.0 - cum_descent[j] / total_descent
        if remaining < remaining_threshold:
            return peak_idx + j

    # 全段都是慢下降（饱和型：下降集中在末尾）
    return end_idx


# ── Stage 3b: B通道电阻分析 ────────────────────────────────────────

def analyze_resistance_per_round(
    resistances: np.ndarray,
    start_idx: int,
    peak_idx: int,
    end_idx: int,
    n_points: int = 5,
    pre_window: int = 24,
    post_window: int = 24,
) -> Optional[ResistanceRoundInfo]:
    """
    对单轮实验计算电阻分析指标（通道B：电阻）。

    Rmax: 上升沿前一区间 (start_idx-pre_window, start_idx) 内电阻的稳健最大值
          （取最大的 n_points 个点平均）
    Rmin: 上升沿后区间 (peak_idx, peak_idx+post_window) 内电阻的稳健最小值
          （取最小的 n_points 个点平均）
    区间均以通道A（电流）的分割边界为基准。
    """
    # 上升沿前区间
    pre_start = max(0, start_idx - pre_window)
    pre_rise_R = resistances[pre_start:start_idx]

    # 上升沿后区间
    post_start = peak_idx
    post_end = min(peak_idx + post_window, end_idx)
    if post_end <= post_start:
        post_end = end_idx + 1
    post_rise_R = resistances[post_start:post_end]

    if len(pre_rise_R) == 0 or len(post_rise_R) == 0:
        return None

    r_max = _trimmed_max(pre_rise_R, n_points)
    r_min = _trimmed_min(post_rise_R, n_points)

    if r_min <= 0:
        return None

    return ResistanceRoundInfo(
        r_max=r_max,
        r_min=r_min,
        r_ratio=r_max / r_min,
    )


# ── Stage 4: 管道编排 ──────────────────────────────────────────────

def run_analysis(
    filepath: str,
    rise_lower_percent: float = 10.0,
    rise_upper_percent: float = 90.0,
    fall_upper_percent: float = 90.0,
    fall_lower_percent: float = 10.0,
    prominence: float = 0.3,
    distance: int = 150,
    baseline_fraction: float = 0.01,
    peak_threshold: float = 1e-7,
    b_channel: str = "",
) -> AnalysisResult:
    """
    运行完整的分析管道。

    Stages:
      0. 加载数据
      1. 平滑电流
      2. 实验阶段分割（基于电流峰值检测）
      3. 对每轮提取电流上升/下降边沿（通道A）
      4. （可选）B通道分析：电阻 → 每轮Rmax/Rmin；电压 → 全局RMS/RMSE

    参数:
        b_channel: "" (无), "resistance", "voltage"
    """
    # Stage 0: 加载
    times, voltages, currents, resistances = load_data(filepath)

    # Stage 1: 平滑电流
    currents_smooth = smooth_current(currents, window=21, poly=3)

    # Stage 2: 分割
    round_boundaries = find_round_boundaries(
        times, currents_smooth,
        prominence=prominence,
        distance=distance,
        baseline_fraction=baseline_fraction,
        peak_threshold=peak_threshold,
    )

    # Stage 3: 通道A边沿提取
    rounds = []
    for i, (start, peak, end) in enumerate(round_boundaries):
        rise_info = analyze_edge(
            times, currents, currents_smooth, start, peak, is_rising=True,
            lower_percent=rise_lower_percent,
            upper_percent=rise_upper_percent,
        )

        # 峰值类型判断
        is_overshoot = detect_peak_type(currents_smooth, peak, end, start_idx=start)

        overshoot_info = None
        if is_overshoot:
            # 找快速下降起始点
            rapid_start = find_rapid_descent_start(currents_smooth, peak, end)

            # 计算 I_90_Stable：过冲下降结束的稳定电流值
            # 公式：(过冲峰值 - 气体结束前电流) * (100%-Y%) + 气体结束前电流
            # 其中 Y% = rise_upper_percent
            peak_val = currents_smooth[peak]
            rapid_start_val = currents_smooth[rapid_start]
            y_frac = rise_upper_percent / 100.0
            i_90_stable = (peak_val - rapid_start_val) * (1.0 - y_frac) + rapid_start_val

            # 找慢下降中电流达到I_90_Stable的交叉时间
            slow_descent_time = 0.0
            overshoot_descent_time = float(times[rapid_start]) - float(times[start])
            if rapid_start > peak:
                slow_seg = currents_smooth[peak:rapid_start + 1]
                slow_times = times[peak:rapid_start + 1]
                # 从峰顶向右扫描，找电流首次降到 I_90_Stable 以下
                for j in range(len(slow_seg)):
                    if slow_seg[j] <= i_90_stable:
                        if j > 0 and slow_seg[j - 1] > i_90_stable:
                            t1, t2 = slow_times[j - 1], slow_times[j]
                            i1, i2 = slow_seg[j - 1], slow_seg[j]
                            frac = (i1 - i_90_stable) / (i1 - i2) if i1 != i2 else 0
                            cross_time = t1 + frac * (t2 - t1)
                        else:
                            cross_time = slow_times[j]
                        slow_descent_time = float(cross_time) - float(times[peak])
                        overshoot_descent_time = float(cross_time) - float(times[start])
                        break

            # 快速下降边沿分析（气体通入结束的快速下降）
            fall_info = analyze_edge(
                times, currents, currents_smooth,
                rapid_start, end, is_rising=False,
                lower_percent=fall_lower_percent,
                upper_percent=fall_upper_percent,
            )

            overshoot_info = OvershootInfo(
                rapid_descent_start_idx=rapid_start,
                rapid_descent_start_time=float(times[rapid_start]),
                i_90_stable=i_90_stable,
                slow_descent_time=slow_descent_time,
                overshoot_descent_time=overshoot_descent_time,
            )
        else:
            # 饱和峰 — 保持原有逻辑
            fall_info = analyze_edge(
                times, currents, currents_smooth, peak, end, is_rising=False,
                lower_percent=fall_lower_percent,
                upper_percent=fall_upper_percent,
            )

        # Stage 3b: B通道电阻分析（按需，基于通道A分割边界）
        res_info = None
        if b_channel == "resistance":
            res_info = analyze_resistance_per_round(
                resistances, start, peak, end,
                n_points=5, pre_window=24, post_window=24,
            )

        rounds.append(RoundResult(
            round_num=i + 1,
            peak_time=times[peak],
            peak_current=currents_smooth[peak],
            rise=rise_info,
            fall=fall_info,
            start_idx=start,
            peak_idx=peak,
            end_idx=end,
            res_info=res_info,
            is_overshoot=is_overshoot,
            overshoot_info=overshoot_info,
        ))

    return AnalysisResult(
        times=times,
        voltages=voltages,
        currents=currents,
        resistances=resistances,
        currents_smooth=currents_smooth,
        rounds=rounds,
        file_path=filepath,
    )
